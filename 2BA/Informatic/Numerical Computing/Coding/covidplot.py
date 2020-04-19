# covidplot.py
# Author: Mitrovic Nikola
# Version: April 19, 2020

import openpyxl, os
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import datetime as dt

ROOT = os.path.abspath(os.getcwd()) + "/2BA/Informatic/Numerical Computing/Coding/"

# Opening COVID19 dataset in xlsx format
# Select sheet DEATH
book = openpyxl.load_workbook(ROOT + 'COVID19BE.xlsx')
sheet = book.get_sheet_by_name('MORT')

# Initialize some variables 
# For the data analysis below
data = []
death = 0
i = 2
max_row = sheet.max_row + 1

# Retrieve date and death in dataset
# Cumulute death per day and store data
while i < max_row:
    death += sheet.cell(row = i, column = 5).value
    date = sheet.cell(row = i, column = 1).value
    if date == sheet.cell(row = i+1, column = 1).value :
        death += sheet.cell(row = i+1, column = 5).value
    if date != sheet.cell(row = i+1, column = 1).value :
        data.append([date, death])
        death = 0
    i+=1

# Retrieve date in current format
# Separte year, month and day and store into 'x' values
# And finally set y values from data    
dates = [date[0] for date in data]
x = [dt.datetime.strptime(d, '%Y-%m-%d').date() for d in dates]
y = [death[1] for death in data]

# Set the figure and axes
# Create axes label, plot title and grid
fig, ax = plt.subplots()
ax.set_xlabel('Time [MM-DD]', fontsize='12')
ax.set_ylabel('Deaths', fontsize='12')
ax.set_title('Total COVID-19 Deaths in Belgium 2020', fontsize='16', fontweight='bold')
ax.grid()

# Convert the date in 'MM-DD' format for the 'x' axe
# Plot using line with dot in red color and show the plot
plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%m-%d'))
plt.gca().xaxis.set_major_locator(mdates.DayLocator(interval=4))
plt.plot(x,y, 'o-r', linewidth=2)
plt.gcf().autofmt_xdate()
plt.show()