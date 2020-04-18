# dataset.py
# Author: Mitrovic Nikola
# Version: April 16, 2020

import openpyxl, os, datetime
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.ticker.formatter as formatter
import matplotlib.ticker.locator as locator
import numpy as np

ROOT = os.path.abspath(os.getcwd()) + "/2BA/Informatic/Numerical Computing/Coding/"

# Opening COVID19 dataset in xlsx format:
# Plot data analyses

book = openpyxl.load_workbook(ROOT + 'COVID19BE.xlsx')
sheet = book.get_sheet_by_name('MORT')
shape = (sheet.max_row, 4)

data = []
death = 0
i = 2

while i < sheet.max_row :
    death += sheet.cell(row = i, column = 5).value
    date = sheet.cell(row = i, column = 1).value
    if date == sheet.cell(row = i+1, column = 1).value :
        death += sheet.cell(row = i+1, column = 5).value
    if date != sheet.cell(row = i+1, column = 1).value :
        data.append([date, death])
        death = 0
    i+=1

# fig, ax = plt.subplots(figsize=(12, 12))
dates = []
for i in range(len(data)):
    dates.append(data[i][0])
x_values = [datetime.datetime.strptime(d, "%Y-%m-%d").date() for d in dates]
y_values = [item[1] for item in data]
ax = plt.gca()
formatter = mdates.DateFormatter("%m-%d")
ax.xaxis.set_major_formatter(formatter)
locator = mdates.DayLocator()
ax.xaxis.set_major_formatter(locator)
plt.plot(x_values, y_values[:])

# date_form = ("%m-%d")
# fig, ax = plt.subplot(figsize=())
# ax.xaxis.set_major_formatter(date_form)
# plt.plot(data[:,2], data[:, 3])
# plt.grid()
# plt.show()